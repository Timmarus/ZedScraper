from openpyxl import Workbook, load_workbook
import random


def get_percentage(percentage_order):
    list = [1]*percentage_order[0]+\
           [2]*percentage_order[1]+\
           [3]*percentage_order[2]+\
           [4]*percentage_order[3]+\
           [5]*percentage_order[4]
    return random.choice(list)

def run(gender, height_max, height_min, interests, age_max, age_min, weight_max, weight_min, percentage_order):
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def excel_style(row, col):
        """ Convert given row and column number to an Excel-style cell name. """
        result = []
        while col:
            col, rem = divmod(col - 1, 26)
            result[:0] = LETTERS[rem]
        return ''.join(result) + str(row)

    wb = load_workbook("default.xlsx")
    ws = wb.active
    count = 2
    city_file = open("city codes.txt").readlines()
    interests_file = open(interests).readlines()
    for row in ws.rows:
        interest_len = random.randint(3, 11)
        interests = ""
        for i in range(0, interest_len):
            interests += random.choice(interests_file) + ", "
        interests = interests[:-2]
        #print(ws[excel_style(count, 1)].value)
        ws[excel_style(count, 2)].value = str(count)+"@abc.com"
        ws[excel_style(count, 3)].value = gender
        ws[excel_style(count, 5)].value = random.randint(age_min, age_max)
        ws[excel_style(count, 10)].value = 180
        ws[excel_style(count, 11)].value = random.choice(city_file)
        ws[excel_style(count, 21)].value = get_percentage(percentage_order)
        ws[excel_style(count, 20)].value = interests
        ws[excel_style(count, 24)].value = random.randint(height_min, height_max)
        ws[excel_style(count, 25)].value = random.randint(weight_min, weight_max)
        count+=1
    wb.save("default.xlsx")

if __name__ == "__main__":

    gender = "female"

    height_max = 175
    height_min = 157

    interest_file = "interests.txt"

    age_max = 37
    age_min = 25

    weight_max = 59
    weight_min = 49

    percentage_order = [0, 50, 10, 20, 20] #Order of percentages for package ids 1, 2, 3, 4, 5

    run(gender, height_max, height_min, interest_file, age_max, age_min, weight_max, weight_min, percentage_order)
