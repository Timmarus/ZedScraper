import requests, http.cookiejar
from bs4 import BeautifulSoup as BS
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import random

def load_cookies(site):
    cookies = []
    f = open("cookies.txt", "r")
    for line in f:
        if site in line:
            text = line.split()
            domain = text[0]
            name = text[5]
            data = text[6]
            tup = (name, data, domain)
            cookies.append(tup)
    return cookies

def victoria_milan(phantom_js_path, gender="female", login_name="Marcello9438", login_pass="koprewpokk432432gfD"):

    scrape_country = "Italy"
    myProxy = '83.136.105.193:26727'

    service_args = [
        '--proxy='+myProxy,
        '--proxy-type=http',
        '--proxy-auth=utorests:D2VVWR086MI3AASWTANRJXIT'
    ]
    binary = phantom_js_path
    if myProxy == "":
        driver = webdriver.PhantomJS(executable_path=binary)
    else:
        driver = webdriver.PhantomJS(executable_path=binary, service_args=service_args)
    driver.set_window_size(1366,768)
    driver.delete_all_cookies()
    print("Beginning login sequence.")
    driver.get('https://www.victoriamilan.co.uk/login')
    driver.save_screenshot('file.png')
    username, password = driver.find_element_by_id("username"), driver.find_element_by_id("password")
    username.send_keys(login_name)
    password.send_keys(login_pass)
    password.send_keys(Keys.RETURN)
    ids, done, count, total = [], False, 1, 0
    while not done:
        if count >= 2:
            break
        driver.save_screenshot('screen.png')
        if gender == "female":
            url = 'https://www.victoriamilan.co.uk/search/?search[filter]=' \
                  '&search[username]=&search[city]=3169070&search[distanceStep]=6' \
                  '&search[seekingSex]=1&search[ageFrom]=18&search[ageTo]=100&page=' + str(count)
        elif gender == "male":
            url = 'https://www.victoriamilan.co.uk/search/?search[filter]=' \
                  '&search[username]=&search[city]=3169070&search[distanceStep]=6' \
                  '&search[seekingSex]=2&search[ageFrom]=18&search[ageTo]=100&page=' + str(count)
        driver.get(url)
        driver.save_screenshot('screen2.png')
        wait = WebDriverWait(driver, 3)
        try:
            searchResults = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "sr-profile-card")))
        except:
            print("Empty page...")
        html = driver.page_source
        soup = BS(html, 'html.parser')
        divs = soup.find_all("div")
        if 'soup' in locals() and not soup.find("span", {"class": "count"}):
            done = True
            print("Finished parsing IDs.")
            break
        for i in divs:
            if 'data-user-id' in i.attrs:
                ids.append(i.attrs['data-user-id'])
        print("Page " + str(count) + " done.")
        count+=1
    print("Attempting to scrape from " +str(len(ids))+" profiles.")
    count = 1

    for i in ids:
        with open("db.txt") as file:
            if str(i) in file.read():
                print("Already in file.")
                count+=1
                continue
        driver.get("https://www.victoriamilan.co.uk/profile/"+str(i))
        driver.save_screenshot("profile.png")
        try:
            name = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "user-name")))
        except:
            count+=1
            continue
        html = driver.page_source
        if scrape_country in html and "name-box" in html:
            soup = BS(html, 'html.parser')
            country = soup.find("span", {'class': 'location-label'})
            if scrape_country in country.text:
                name = soup.find("div", {'class': 'name-box'}).find("span", {'class': 'user-name'}).text
                bio = soup.find("p", {"data-field": "about_you"})
                if bio == "" or bio is None:
                    bio = str(None)
                else:
                    bio = bio.text
                print(str(i) + " successfully scraped. " + str((len(ids) - count)) + " remaining.")
                driver.save_screenshot("last_succeed.png")
                write_data(name, bio, i, "VictoriaMilan")
                total += 1
            else:
                print("Italy not found.")
        else:
            print(str(i) + " invalid.")
        count += 1
    driver.quit()
    print(str(total) + " profiles successfully scraped.")

def sexytribute(phantom_js_path, limit=50, gender="female"):
    cj = http.cookiejar.MozillaCookieJar('cookies.txt')
    cj.load()
    count, done, ids, last_page, profiles = 1, False, [], 0, {}
    page_limit = limit
    print("Scraping usernames from " + str(page_limit) + " pages.")
    myProxy = '83.136.105.193:26727'
    service_args = [
        '--proxy='+myProxy,
        '--proxy-type=http',
        '--proxy-auth=utorests:D2VVWR086MI3AASWTANRJXIT'
    ]
    binary = phantom_js_path
    driver = webdriver.PhantomJS(executable_path=binary, service_args=service_args)
    driver.get("http://www.sexytribu.it/login")
    driver.save_screenshot("test.png")
    username, password = driver.find_element_by_name("login"), driver.find_element_by_name("password")
    username.send_keys("marcellomule@hotmail.com")
    password.send_keys("password")
    password.send_keys(Keys.RETURN)
    wait = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CLASS_NAME, "MembersList")))
    driver.get("http://www.sexytribu.it/search/advanced/?action=searchAction&mySexId=1&seekingSexId=2&seekingAgeFrom=18&seekingAgeTo=90&seekingCountryId=100&seekingRegionId=0&usePostalCode=false")
    wait = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CLASS_NAME, "page")))
    req = driver.page_source
    srchN = req.split("srchN=")[1].split("&")[0]
    driver.quit()
    while count <= page_limit:
        if gender == "female":
            url = 'http://www.sexytribu.it/search/advanced/?pageNum=' + str(
                count) + '&action=searchAction&mySexId=1&onlineOnly=false' \
                         '&usePostalCode=false&postalCode=&seekingAgeFrom=18' \
                         '&seekingAgeTo=90&seekingCountryId=100&seekingOrientationId=0' \
                         '&seekingRegionId=0&seekingSexId=2&srchN='+srchN
        elif gender == "male":
            url = 'http://www.sexytribu.it/search/advanced/?pageNum=' + str(
                count) + '&action=searchAction&mySexId=1&onlineOnly=false' \
                         '&usePostalCode=false&postalCode=&seekingAgeFrom=18' \
                         '&seekingAgeTo=90&seekingCountryId=100&seekingOrientationId=0' \
                         '&seekingRegionId=0&seekingSexId=1&srchN='+srchN

        r = requests.get(url, cookies=cj)
        html = r.text
        soup = BS(html, "html.parser")
        members = soup.find_all("div", {"class": "member"})
        page = soup.find_all("span", {"class": "page"})[-1].text
        for member in members:
            ids.append(member.find("img")['id'])

        if page == last_page:
            print("Finished scraping usernames from", str(page_limit), "pages.")
            break
        if int(page) % 5 == 0:
            print(str(page_limit-count) + " pages remaining.")
        last_page = page
        count+=1

    print("Attempting to scrape " + str(len(ids)) + " profiles.")
    count = 0

    for id in ids:
        with open("db.txt") as file:
            if str(id) in file.read():
                print("Already in file.")
                count+=1
                continue
        url = "http://sexytribu.it/members/" + str(id)
        r = requests.get(url, cookies=cj)
        html = r.text
        soup = BS(html, "html.parser")
        try:
            name = soup.find("div", {"class": "ProfileMenu"}).find("h1").find("strong").text
        except:
            continue
        try:
            bio = soup.find("div", {"class": "google-translate-text"}).text
        except:
            bio = str(None)
        profiles[name] = (bio, id, "SexyTribute")
        print(str(id) + " scraped. " + str((len(ids) - count)) + " remaining.")
        write_data(name, bio, id, "SexyTribute")
        count +=1


def write_data(name, bio, id, site):
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    def excel_style(row, col):
        """ Convert given row and column number to an Excel-style cell name. """
        result = []
        while col:
            col, rem = divmod(col - 1, 26)
            result[:0] = LETTERS[rem]
        return ''.join(result) + str(row)

    with open("db.txt", "a+") as file:
        if str(id) in file.read():
            print("Already in file.")
            return
        try:
            file.write("################################################\n")
            file.write("Name: " + name +
                       "\nBio: " + str(bio) +
                       "\n" + "ID: " + str(id) +
                       "\n" + "Site: + " + site + "\n")
            file.write("################################################\n")
        except:
            pass

    wb = load_workbook("default.xlsx")
    ws = wb.get_sheet_by_name("Sheet1")
    row_count = 1
    for i in range(0, ws.max_row+10):
        if ws.cell(row=row_count, column=1).value is None:
            break
        row_count +=1
    new_row = row_count
    alment = Alignment(wrap_text=True)
    ws[excel_style(new_row, 1)].alignment,\
    ws[excel_style(new_row, 15)].alignment = alment, alment
    ws[excel_style(new_row, 1)].value = name
    ws[excel_style(new_row, 15)].value = bio
    new_row+=1
    wb.save("default.xlsx")

def vk(phantom_js_path):
    def get_name_and_row():
        wb = load_workbook("default.xlsx")
        ws = wb.active
        i = 1
        for row in range(0, ws.max_row +5):
            if ws.cell(row=i, column=30).value in [None, 'None']:
                return (ws.cell(row=i, column=1).value, i)
            i+=1
    def save_imgs(imgs):
        try:
            info = get_name_and_row()
            name = info[0]
            row = info[1]
            wb = load_workbook("default.xlsx")
            ws = wb.active
            count = 1
            num_img = random.randint(1, 5)
            completed = []
            for img in imgs:
                if img in completed:
                    print("continued")
                    continue
                r = requests.get(img)
                extension = img[-4:]
                filename = str(name)+str(count)+extension
                open("imgs/" + filename, "wb").write(r.content)
                if count <= num_img:
                    if count == 1:
                        ws.cell(row=row, column=17).value = filename
                    pre = ws.cell(row=row, column=30).value
                    if pre == None:
                        ws.cell(row=row, column=30).value =  filename
                    else:
                        ws.cell(row=row, column=30).value = str(pre) + ", " + filename
                else:
                    pre = ws.cell(row=row, column=31).value
                    if pre == None:
                        ws.cell(row=row, column=31).value = filename
                    else:
                        ws.cell(row=row, column=31).value = str(pre) + ", " + filename
                count+=1
                completed.append(img)
                if count >= 10:
                    break
            wb.save("default.xlsx")
        except:
            print("Returned.")
            return


    binary = phantom_js_path
    driver = webdriver.PhantomJS(executable_path=binary)
    driver.set_window_size(1366,768)
    driver.maximize_window()
    file = open("vk.txt", "r").readlines()
    for line in file:
        if line.isspace():
            continue
        url = line.rstrip()
        print(url)
        imgs = []

        driver.get(url)
        if "Only logged in users can see this profile." in driver.page_source:
            print(url, "failed.")
            continue
        try:
            wait = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CLASS_NAME, "page_square_photo")))
        except:
            print("Continued.")
            continue
        driver.save_screenshot('vk.png')
        soup = BS(driver.page_source, "html.parser")
        href = soup.find("a", {"class": "crisp_image"})['href']
        driver.get('http://vk.com/' + href)
        for i in range(0, 9):
            try:
                wait = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "pv_img_area_wrap")))
            except:
                print("Continued 2.")
                continue
            soup = BS(driver.page_source, "html.parser")
            imgs.append(soup.find("div", {"class": "pv_img_area_wrap"}).find("img")['src'])
            try:
                driver.find_element_by_id("pv_photo").click()
            except:
                print("Continued 3.")
                continue
        save_imgs(imgs)
    driver.quit()
    print("Done.")


if __name__ == "__main__":

    phantom_js_path = "/Users/admin/Downloads/phantomjs-2.1.1-windows/phantomjs-2.1.1-windows/bin/phantomjs.exe"

    gender = "female" #Set to either "male" or "female"

    login_name = "" #Username for victoriamilan login
    login_pass = "" #Password for victoriamilan login

    # sexytribute(phantom_js_path, 1, gender)
    # victoria_milan(phantom_js_path, gender) #Comment this to use your own login and password
    #victoria_milan(gender, login_name, login_pass) #Uncomment this to use your preset username and password for victoriamilan
    vk(phantom_js_path)