from openpyxl import Workbook, load_workbook
import os


def gall(ws, original, changed, row):
    pre = ws.cell(row=row, column=30).value.replace(original, changed)
    ws.cell(row=row, column=30).value = pre
    pre = ws.cell(row=row, column=31).value.replace(original, changed)
    ws.cell(row=row, column=31).value = pre


def name_change(original, changed):
    wb = load_workbook("default.xlsx")
    ws = wb.active
    row_count = 1
    for row in ws.rows:
        if ws.cell(row=row_count, column=1).value == original:
            break
        row_count+=1
    if row_count > 1:
        ws.cell(row=row_count, column=1).value = changed
        gall(ws, original, changed, row_count)
        fnames = os.listdir("imgs")
        for fname in fnames:
            if fname.startswith(original):
                os.rename("imgs/"+fname, "imgs/"+fname.replace(original, changed))
        wb.save("default.xlsx")
    else:
        raise Exception("Cannot change first row.")



if __name__ == "__main__":
    original = "Zarinaas" #Original Name
    changed = "Zarinas" #Changed name

    name_change(original, changed)