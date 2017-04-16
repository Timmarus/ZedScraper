import requests, http.cookiejar
from urllib.parse import quote
from bs4 import BeautifulSoup as BS
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.proxy import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from random import randint
import random

# def vk():
#     def get_name_and_row():
#         wb = load_workbook("default.xlsx")
#         ws = wb.active
#         i = 1
#         for row in ws.rows:
#             if ws.cell(row=i, column=30).value in [None, 'None']:
#                 return (ws.cell(row=i, column=1).value, i)
#             i+=1
#     def save_imgs(imgs):
#         try:
#             info = get_name_and_row()
#             name = info[0]
#             row = info[1]
#             wb = load_workbook("default.xlsx")
#             ws = wb.active
#             count = 1
#             num_img = randint(1, 5)
#             lastimg = None
#             for img in imgs:
#                 if lastimg == img:
#                     continue
#                 lastimg = img
#                 r = requests.get(img)
#                 extension = img[-4:]
#                 filename = str(name+str(count)+extension)
#                 open("imgs/" + filename, "wb").write(r.content)
#                 if count <= num_img:
#                     pre = ws.cell(row=row, column=30).value
#                     if pre == None:
#                         ws.cell(row=row, column=30).value =  filename
#                     else:
#                         ws.cell(row=row, column=30).value = str(pre) + ", " + filename
#                 else:
#                     pre = ws.cell(row=row, column=31).value
#                     if pre == None:
#                         ws.cell(row=row, column=31).value = filename
#                     else:
#                         ws.cell(row=row, column=31).value = str(pre) + ", " + filename
#                 count+=1
#                 if count >= 10:
#                     break
#             wb.save("default.xlsx")
#         except:
#             return
#
#
#     binary = "/Users/admin/Downloads/phantomjs-2.1.1-windows/phantomjs-2.1.1-windows/bin/phantomjs.exe"
#     driver = webdriver.PhantomJS(executable_path=binary)
#     #driver = webdriver.Chrome("chromedriver.exe")
#     driver.set_window_size(1366,768)
#     driver.maximize_window()
#     file = open("vk.txt", "r").readlines()
#     for line in file:
#         if line.isspace():
#             continue
#         url = line.rstrip()
#         print(url)
#         imgs = []
#
#         driver.get(url)
#         if "Only logged in users can see this profile." in driver.page_source:
#             print(url, "failed.")
#             continue
#         wait = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CLASS_NAME, "page_square_photo")))
#         driver.save_screenshot('vk.png')
#         soup = BS(driver.page_source, "html.parser")
#         href = soup.find("a", {"class": "crisp_image"})['href']
#         driver.get('http://vk.com/' + href)
#         for i in range(0, 9):
#             wait = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "pv_img_area_wrap")))
#             soup = BS(driver.page_source, "html.parser")
#             imgs.append(soup.find("div", {"class": "pv_img_area_wrap"}).find("img")['src'])
#             #driver.find_element_by_xpath("//img[@src='"+imgs[0]+"']").click()
#             try:
#                 driver.find_element_by_id("pv_photo").click()
#             except: continue
#             #driver.save_screenshot('vk.png')
#         save_imgs(imgs)
#
#     print("Done.")
myProxy = '83.136.105.193:26727'
service_args = [
    '--proxy=' + myProxy,
    '--proxy-type=http',
    '--proxy-auth=utorests:D2VVWR086MI3AASWTANRJXIT'
]

username="vieru.andrei90@yahoo.com"
password="okp#%POdAFPO#%#sdfdsf34"
phantom_js_path = "/Users/admin/Downloads/phantomjs-2.1.1-windows/phantomjs-2.1.1-windows/bin/phantomjs.exe"
binary = phantom_js_path
driver = webdriver.PhantomJS(executable_path=binary, service_args=service_args)
driver.set_window_size(1366, 768)
driver.maximize_window()
driver.get("https://vk.com/")
wait = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.ID, "email")))
driver.save_screenshot("test.png")
user, passw = driver.find_element_by_id("index_email"), driver.find_element_by_id("index_pass ")
user.send_keys(username)
passw.send_keys(password)
user.send_keys(Keys.RETURN)
time.sleep(5)
wait = WebDriverWait(driver, 8)
driver.save_screenshot("test.png")


def vk(phantom_js_path, username="vieru.andrei90@yahoo.com", password="okp#%POdAFPO#%#sdfdsf34"):
    def get_name_and_row():
        wb = load_workbook("default.xlsx")
        ws = wb.active
        i = 1
        for row in ws.rows:
            if ws.cell(row=i, column=30).value in [None, 'None']:
                return (ws.cell(row=i, column=1).value, i)
            i+=1
    def save_imgs(imgs):
        try:
            info = get_name_and_row()
            name = info[0]
            row = info[1]
            wb = load_workbook("default.xlsx")
            ws = wb.active
            count = 1
            num_img = random.randint(1, 5)
            lastimg = None
            for img in imgs:
                if lastimg == img:
                    continue
                lastimg = img
                r = requests.get(img)
                extension = img[-4:]
                filename = str(name+str(count)+extension)
                open("imgs/" + filename, "wb").write(r.content)
                if count <= num_img:
                    pre = ws.cell(row=row, column=30).value
                    if pre == None:
                        ws.cell(row=row, column=30).value =  filename
                    else:
                        ws.cell(row=row, column=30).value = str(pre) + ", " + filename
                else:
                    pre = ws.cell(row=row, column=31).value
                    if pre == None:
                        ws.cell(row=row, column=31).value = filename
                    else:
                        ws.cell(row=row, column=31).value = str(pre) + ", " + filename
                count+=1
                if count >= 10:
                    break
            wb.save("default.xlsx")
        except:
            return


    binary = phantom_js_path
    driver = webdriver.PhantomJS(executable_path=binary)
    driver.set_window_size(1366,768)
    driver.maximize_window()
    file = open("vk.txt", "r").readlines()
    for line in file:
        if line.isspace():
            continue
        url = line.rstrip()
        print(url)
        imgs = []

        driver.get(url)
        if "Only logged in users can see this profile." in driver.page_source:
            print(url, "failed.")
            continue
        try:
            wait = WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.CLASS_NAME, "page_square_photo")))
        except:
            continue
        driver.save_screenshot('vk.png')
        soup = BS(driver.page_source, "html.parser")
        href = soup.find("a", {"class": "crisp_image"})['href']
        driver.get('http://vk.com/' + href)
        for i in range(0, 9):
            try:
                wait = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "pv_img_area_wrap")))
            except:
                continue
            soup = BS(driver.page_source, "html.parser")
            imgs.append(soup.find("div", {"class": "pv_img_area_wrap"}).find("img")['src'])
            try:
                driver.find_element_by_id("pv_photo").click()
            except: continue
        save_imgs(imgs)
    driver.quit()
    print("Done.")
